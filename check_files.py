import os
from openpyxl import load_workbook

def check_excel_file(file_path, name):
    """Проверяет содержимое Excel файла"""
    print(f"\n=== ПРОВЕРКА {name} ===")
    print(f"Файл: {file_path}")
    print(f"Размер: {os.path.getsize(file_path)} байт")

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        print(f"Количество листов: {len(wb.sheetnames)}")
        print(f"Названия листов: {wb.sheetnames}")

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\nЛист '{sheet_name}':")
            print(f"  Размеры: {ws.max_row} строк x {ws.max_column} колонок")

            # Читаем заголовок
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            print(f"  Заголовок: {header_row}")

            # Читаем первые 5 строк данных
            data_rows = list(ws.iter_rows(min_row=2, max_row=6, values_only=True))
            print("  Данные (строки 2-6):")
            for i, row in enumerate(data_rows, 2):
                if any(cell is not None for cell in row):
                    print(f"    Строка {i}: {row}")
                else:
                    print(f"    Строка {i}: пустая")

        wb.close()
        return True

    except Exception as e:
        print(f"Ошибка при чтении файла: {e}")
        return False

# Проверяем файлы
check_excel_file("out/skipped.xlsx", "SKIPPED.XLSX")
check_excel_file("out/errors.xlsx", "ERRORS.XLSX")
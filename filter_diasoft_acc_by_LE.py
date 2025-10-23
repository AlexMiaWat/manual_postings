#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
filter_diasoft_acc_by_LE.py
Фильтрует строки из xlsx-файлов по списку LE (LE.txt), сохраняет
отфильтрованные строки в отдельные xlsx-файлы (в папку out/),
а пропущенные строки собирает в skipped.xlsx. Ошибки логируются в errors.xlsx.
Сохраняет стили ячеек (шрифт, границы, фон, выравнивание и т.д.) и заголовки.
Также удаляет пустые строки в конце выходного листа.
"""

import os
import glob
import re
import datetime
import logging
import pandas as pd
from openpyxl import load_workbook, Workbook
from pathlib import Path
from copy import copy
try:
    from rich.console import Console
    from rich.text import Text
    RICH_AVAILABLE = True
except ImportError:
    RICH_AVAILABLE = False

# ========== Настройки файлов и директорий ==========
LE_FILE = "LE.txt"
OUT_DIR = "out"
SKIPPED_FILE = "skipped.xlsx"
ERRORS_FILE = "errors.xlsx"
LOG_FILE = "log.md"  # также используем logging модуль для файла .md

# ========== Настройки красивого вывода ==========
if RICH_AVAILABLE:
    console = Console()
else:
    console = None

# Создаём папку out если нет
os.makedirs(OUT_DIR, exist_ok=True)

# Очищаем папку out (удаляем файлы внутри)
for f in os.listdir(OUT_DIR):
    try:
        os.remove(os.path.join(OUT_DIR, f))
    except Exception:
        pass

# ========== Настройка логирования ==========
logger = logging.getLogger("filter_le")
logger.setLevel(logging.DEBUG)

# File handler - пишет в LOG_FILE (append) чистый MD без префиксов
fh = logging.FileHandler(LOG_FILE, encoding="utf-8")
fh.setLevel(logging.DEBUG)
fh_formatter = logging.Formatter("%(message)s")  # Только сообщение, без времени и уровня
fh.setFormatter(fh_formatter)

# Console handler - INFO и выше в консоль с цветами если rich доступен
ch = logging.StreamHandler()
ch.setLevel(logging.INFO)
ch_formatter = logging.Formatter("%(levelname)s: %(message)s")
ch.setFormatter(ch_formatter)

logger.addHandler(fh)
logger.addHandler(ch)

# ========== Вспомогательные функции для красивого вывода ==========

from rich.table import Table
from rich.panel import Panel
from rich.markdown import Markdown
from datetime import datetime

def log_md(message: str, level: str = "INFO"):
    timestamp = datetime.now().strftime("%H:%M:%S")
    prefix = {
        "INFO": "✅",
        "WARNING": "⚠️",
        "ERROR": "❌",
    }.get(level, "ℹ️")
    md_message = f"> **[{timestamp}] {prefix} {message}**"
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(f"{md_message}\n")

    # Красивый вывод в консоль
    if console:
        color_map = {"INFO": "green", "WARNING": "yellow", "ERROR": "red"}
        color = color_map.get(level, "white")
        console.print(Panel(f"[{color}]{message}[/]", title=level, expand=False))
    else:
        print(f"{timestamp} {level}: {message}")

def log_header(text: str, level: int = 1):
    """
    Логирует заголовок MD.
    """
    hashes = "#" * level
    msg = f"{hashes} {text}"
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(f"{msg}\n")

def log_list_item(text: str):
    """
    Логирует элемент списка MD.
    """
    msg = f"- {text}"
    log_md(msg, "INFO")

def log_table(headers: list, rows: list):
    """
    Логирует таблицу MD.
    """
    if not headers or not rows:
        return
    # Заголовок
    header_line = "| " + " | ".join(headers) + " |"
    separator = "| " + " | ".join(["---"] * len(headers)) + " |"
    table_lines = [header_line, separator]
    for row in rows:
        row_line = "| " + " | ".join(str(cell) for cell in row) + " |"
        table_lines.append(row_line)
    msg = "\n".join(table_lines)
    log_md(msg, "INFO")

def log_file_separator():
    """
    Добавляет визуальный разделитель между обработкой файлов в консоли.
    """
    if console:
        console.print("\n" + "=" * 60 + "\n", style="blue")
    else:
        print("\n" + "=" * 60 + "\n")

def log_total_stats(filtered: int, skipped: int, errors: int):
    """
    Логирует итоговую статистику в красивом формате.
    """
    if console:
        try:
            console.print("\n[bold cyan]ИТОГОВАЯ СТАТИСТИКА:[/bold cyan]")
            console.print(f"   [green]Отфильтровано: {filtered} строк[/green]")
            console.print(f"   [yellow]Пропущено: {skipped} строк[/yellow]")
            console.print(f"   [red]Ошибок: {errors} строк[/red]")
            console.print("[blue]" + "=" * 60 + "[/blue]")
        except UnicodeEncodeError:
            console.print("\n[bold cyan]ИТОГОВАЯ СТАТИСТИКА:[/bold cyan]")
            console.print(f"   [green]Отфильтровано: {filtered} строк[/green]")
            console.print(f"   [yellow]Пропущено: {skipped} строк[/yellow]")
            console.print(f"   [red]Ошибок: {errors} строк[/red]")
            console.print("[blue]" + "=" * 60 + "[/blue]")
    else:
        print(f"\nИТОГОВАЯ СТАТИСТИКА:")
        print(f"Отфильтровано: {filtered} строк")
        print(f"Пропущено: {skipped} строк")
        print(f"Ошибок: {errors} строк")

# Логируем старт скрипта
log_header("Запуск скрипта filter_diasoft_acc_by_LE.py")
log_header("Настройки", 2)
log_header(f"LE файл: {LE_FILE}", 3)
log_header(f"Выходная папка: {OUT_DIR}", 3)
log_header(f"Файл пропущенных: {SKIPPED_FILE}", 3)
log_header(f"Файл ошибок: {ERRORS_FILE}", 3)
log_header(f"Лог файл: {LOG_FILE}", 3)
log_md("", "INFO")  # Пустая строка для разделения абзацев

# ========== Вспомогательные функции ==========

def load_le_set(le_file: str) -> set:
    """
    Загружает LE из текстового файла.
    Очищает: пробелы, тире; переводит в верхний регистр.
    Возвращает set строк.
    """
    le_set = set()
    try:
        with open(le_file, "r", encoding="utf-8") as fh:
            for line in fh:
                cleaned = line.strip().replace("-", "").replace(" ", "").upper()
                if cleaned:
                    le_set.add(cleaned)
        log_header("Загруженные LE", 2)
        log_list_item(f"Загружено **{len(le_set)}** LE из `{le_file}`")
        if le_set:
            log_table(["LE"], [[le] for le in sorted(le_set)])
            log_md("", "INFO")  # Пустая строка для разделения абзацев
    except FileNotFoundError:
        logger.error("Файл %s не найден.", le_file)
    except Exception as e:
        logger.exception("Ошибка при чтении %s: %s", le_file, e)
    return le_set

def parse_and_convert_amount(amount_str: str) -> tuple[float | None, str]:
    """
    Преобразует строку суммы в float.
    Очищает разделители разрядов (заменяет запятые) и проверяет формат.
    Возвращает (float, "") при успехе или (None, описание_ошибки).
    """
    if amount_str is None or (isinstance(amount_str, str) and amount_str.strip() == ""):
        return None, "Пустое значение суммы"
    s = str(amount_str).strip()
    # Убираем тысячи-разделители запятой и пробелы
    cleaned = re.sub(r'[,\s]', '', s)
    # Допускаем точку как разделитель дробной части (427680000.00)
    if not re.match(r'^\d+(\.\d+)?$', cleaned):
        return None, f"Некорректный формат суммы: '{amount_str}' -> cleaned '{cleaned}'"
    try:
        val = float(cleaned)
        return val, ""
    except Exception as e:
        return None, f"Ошибка преобразования в число: {e}"

def copy_cell_style(src_cell, dest_cell):
    """
    Копирует визуальные атрибуты ячейки из src_cell в dest_cell.
    Копируются: font, border, fill, number_format, protection, alignment.
    Обёртка через copy() чтобы избежать общих ссылок на объекты стилей.
    """
    if src_cell is None or not getattr(src_cell, "has_style", False):
        return
    try:
        dest_cell.font = copy(src_cell.font)
        dest_cell.border = copy(src_cell.border)
        dest_cell.fill = copy(src_cell.fill)
        dest_cell.number_format = src_cell.number_format
        dest_cell.protection = copy(src_cell.protection)
        dest_cell.alignment = copy(src_cell.alignment)
    except Exception as e:
        logger.warning("Не удалось полностью скопировать стиль ячейки %s: %s", getattr(src_cell, "coordinate", "?"), e)

def is_row_empty(ws, row_idx) -> bool:
    """
    Проверяет, пустая ли строка row_idx на листе ws.
    Считается пустой, если все ячейки None или пустая строка после strip().
    """
    for cell in ws[row_idx]:
        if cell.value is not None and str(cell.value).strip() != "":
            return False
    return True

def remove_trailing_blank_rows(ws):
    """
    Удаляет пустые строк в конце листа ws.
    Пока последняя строка полностью пустая — удаляет её.
    Это предотвращает появление лишней пустой строки в конце выходного файла.
    """
    try:
        max_row = ws.max_row
        # Пытаемся удалить до тех пор, пока последняя строка пустая и > 1 (не удаляем заголовок)
        while max_row > 1:
            # ws[max_row] возвращает кортеж ячеек в строке
            if all((cell.value is None or str(cell.value).strip() == "") for cell in ws[max_row]):
                ws.delete_rows(max_row, 1)
                max_row -= 1
            else:
                break
    except Exception as e:
        logger.exception("Ошибка при удалении пустых строк: %s", e)

# ========== Основная логика обработки одного файла ==========
def write_filtered_rows(file_path: str, le_set: set, skipped_wb: Workbook, errors_ws):
    """
    Читает xlsx файл (через pandas для логики), создаёт/редактирует выходной xlsx
    (используя openpyxl), записывает отфильтрованные строки с сохранением стилей.
    Возвращает (filtered_count, skipped_count, error_count).
    """
    file_name = Path(file_path).name
    log_md(f"Начинаю обработку файла: **{file_name}**", "INFO")

    # 1) Читаем файл в pandas (header=None, чтобы найти реальную строку заголовка программно)
    try:
        df = pd.read_excel(file_path, header=None)
    except Exception as e:
        logger.exception("Ошибка при чтении %s в pandas: %s", file_name, e)
        errors_ws.append([file_name, "", f"Ошибка чтения файла: {e}"])
        return 0, 0, 1

    if df.empty:
        logger.warning("Файл %s пустой.", file_name)
        errors_ws.append([file_name, "", "Файл пустой"])
        return 0, 0, 1

    # 2) Находим первую непустую строку — считаем её заголовком
    header_row = None
    for idx in df.index:
        if df.loc[idx].notna().any():
            header_row = idx
            break
    if header_row is None:
        logger.error("Не удалось найти заголовок в %s", file_name)
        errors_ws.append([file_name, "", "Не найден заголовок"])
        return 0, 0, 1

    # Устанавливаем имена столбцов и отбрасываем строки выше заголовка
    df.columns = df.loc[header_row]
    df = df.loc[header_row + 1 :].reset_index(drop=True)

    # 3) Загружаем workbook через openpyxl — используем файл как шаблон (стили остаются)
    try:
        wb_out = load_workbook(file_path)
    except Exception as e:
        logger.exception("Ошибка load_workbook для %s: %s", file_name, e)
        errors_ws.append([file_name, "", f"Ошибка открытия файла: {e}"])
        return 0, 0, 1

    ws_out = wb_out.active

    # Сохраняем индекс исходной строки заголовка в листе для копирования стиля заголовка
    src_header_row_index = header_row + 1  # потому что pandas 0-based, excel 1-based

    # 4) Очищаем значения на листе, оставляя стили и структуру (формат ячеек не удаляется)
    for row in ws_out.iter_rows(min_row=2):
        for cell in row:
            cell.value = None

    # 5) Устанавливаем заголовки в первой строке листа и копируем стиль из исходного заголовка
    for col_idx in range(len(df.columns)):
        hdr_cell = ws_out.cell(row=1, column=col_idx + 1)
        # Устанавливаем значение заголовка.
        try:
            hdr_cell.value = df.columns[col_idx]
        except Exception:
            # Защита от merged cells или нестандартных заголовков
            hdr_cell.value = str(df.columns[col_idx])
        # Копируем стиль заголовочной ячейки из исходной позиции
        orig_hdr_cell = ws_out.cell(row=src_header_row_index, column=col_idx + 1)
        copy_cell_style(orig_hdr_cell, hdr_cell)

    # Счётчики
    filtered_count = 0
    skipped_count = 0
    error_count = 0
    skipped_rows_indexes = []

    # Логируем начало построчной обработки
    log_list_item(f"Начинаю построчную обработку данных для {file_name}")

    # 6) Проходим строки pandas DataFrame и решаем, записывать ли их в выходной файл
    for row_idx, row in df.iterrows():
        # Пропускаем полностью пустые строки
        if row.isna().all():
            continue

        match_found = False
        is_error = False
        error_desc = ""

        # Ищем значение "LE" в строке (регистр игнорируем)
        for col_idx, (col_name, value) in enumerate(row.items()):
            cell_value = str(value).strip().upper() if pd.notna(value) else ""
            if cell_value == "LE":
                # Берём следующую колонку как "Аналитику"
                if col_idx + 1 < len(row):
                    analytics_raw = row.iloc[col_idx + 1]
                    analytics_value = (
                        str(analytics_raw).strip().replace("-", "").replace(" ", "").upper()
                        if pd.notna(analytics_raw)
                        else ""
                    )
                    if analytics_value and analytics_value in le_set:
                        match_found = True
                        break
                    elif analytics_value:
                        # Не совпало с LE.txt — это нормальная причина пропуска
                        is_error = False
                    else:
                        is_error = True
                        error_desc = f"Пустая аналитика после LE в колонке {col_idx}"
                else:
                    is_error = True
                    error_desc = f"LE в колонке {col_idx}, но нет следующей колонки для аналитики"
                break

        # Если найдено совпадение — записываем строку в выходной лист
        if match_found:
            # Обрабатываем сумму: колонка индекс 7 ожидается числом
            amount_raw = row.iloc[7] if 7 < len(row) else None
            parsed_amount, amt_err = parse_and_convert_amount(amount_raw)
            if amt_err:
                # Записываем в errors_ws и помечаем как ошибка
                row_number = row_idx + src_header_row_index + 1  # для лога: реальный номер в исходном файле
                errors_ws.append([file_name, str(row_number), amt_err])
                logger.warning("Ошибка суммы в %s строка %s: %s", file_name, row_number, amt_err)
                is_error = True
                error_count += 1
                skipped_rows_indexes.append(row_idx)
                skipped_count += 1
                continue
            else:
                # Заменяем значение в pandas-строке на float для корректной записи
                row.iloc[7] = parsed_amount

            # Записываем значения и копируем стили по одной ячейке
            for col_idx in range(len(row)):
                out_cell = ws_out.cell(row=filtered_count + 2, column=col_idx + 1)  # +2: 1 заголовок
                out_cell.value = row.iloc[col_idx]

                # Берём "оригинальную" ячейку из исходного листа (до очистки значений
                # стили всё ещё там). Индекс в исходном листе = header_row + 2 + row_idx
                orig_row_number = src_header_row_index + 1 + row_idx
                orig_cell = ws_out.cell(row=orig_row_number, column=col_idx + 1)
                copy_cell_style(orig_cell, out_cell)

                # Явно назначаем number_format для важных колонок
                if col_idx == 7:
                    out_cell.number_format = "#,##0.00"
                if col_idx == 3:
                    out_cell.number_format = "dd.mm.yyyy"
                    # Попытка преобразовать строку в datetime, если исход был строкой
                    if isinstance(row.iloc[col_idx], str):
                        try:
                            out_cell.value = pd.to_datetime(row.iloc[col_idx])
                        except Exception:
                            # Оставляем как есть и логируем предупреждение
                            logger.debug("Не удалось преобразовать дату '%s' в datetime (файл %s, строка %s)",
                                         row.iloc[col_idx], file_name, filtered_count + src_header_row_index + 1)

            filtered_count += 1

        else:
            # Пропускаем строку — добавляем в skipped список
            skipped_rows_indexes.append(row_idx)
            skipped_count += 1
            if is_error:
                # Если это была ошибка, записываем её
                row_number = row_idx + src_header_row_index + 1
                errors_ws.append([file_name, str(row_number), error_desc])
                error_count += 1
                logger.warning("Строка %s пропущена с ошибкой: %s", row_idx + src_header_row_index + 1, error_desc)

    # 7) Записываем skipped лист если есть пропущенные строки
    if skipped_rows_indexes:
        log_list_item(f"Создаю лист skipped для {file_name}, строк: {len(skipped_rows_indexes)}")
        # Создаём отдельный лист
        skipped_ws = skipped_wb.create_sheet(title=file_name[:31])
        # Копируем заголовок и его стиль из выходного листа (который содержит стиль заголовка)
        for col in range(len(df.columns)):
            cell = skipped_ws.cell(row=1, column=col + 1)
            cell.value = df.columns[col]
            # Копируем стиль из ws_out.header (ячейка 1, col+1)
            copy_cell_style(ws_out.cell(row=1, column=col + 1), cell)

        # Заполняем пропущенные строки значениями и копируем стили из исходного листа
        for i, src_row_idx in enumerate(skipped_rows_indexes):
            row_values = df.iloc[src_row_idx]
            for col_idx in range(len(row_values)):
                val = row_values.iloc[col_idx]
                dest_cell = skipped_ws.cell(row=i + 2, column=col_idx + 1, value=val)
                orig_row_number = src_header_row_index + 1 + src_row_idx
                copy_cell_style(ws_out.cell(row=orig_row_number, column=col_idx + 1), dest_cell)

    # 8) Удаляем пустые строки в конце выходного листа перед сохранением
    remove_trailing_blank_rows(ws_out)

    # 9) Сохраняем выходной файл (только если есть отфильтрованные строки)
    if filtered_count > 0:
        out_file_path = os.path.join(OUT_DIR, file_name)
        try:
            wb_out.save(out_file_path)
            log_list_item(f"Файл сохранён: {out_file_path} (строк: {filtered_count})")
        except Exception as e:
            logger.exception("Ошибка при сохранении %s: %s", out_file_path, e)
            errors_ws.append([file_name, "", f"Ошибка сохранения: {e}"])
            return filtered_count, skipped_count, error_count + 1
    else:
        log_list_item(f"В файле {file_name} нет строк для записи. Выходной файл не создан.")

    # 10) Итоги для файла
    log_md(f"Итог **{file_name}** — Отфильтровано: **{filtered_count}**, Пропущено: **{skipped_count}**, Ошибок: **{error_count}**", "INFO")
    log_md("", "INFO")  # Пустая строка для разделения абзацев

    return filtered_count, skipped_count, error_count

# ========== Точка входа ==========
def main():
    # Логирование старта
    log_header("Старт обработки в папке in/", 2)
    le_set = load_le_set(LE_FILE)
    if not le_set:
        log_md("**Ошибка:** LE список пуст — завершаю.", "ERROR")
        return

    in_files = glob.glob("in/*.xlsx")
    if not in_files:
        log_md("**Ошибка:** Нет файлов в папке `in/`. Завершаю.", "ERROR")
        return

    log_header("Найденные файлы для обработки", 2)
    log_list_item(f"Найдено **{len(in_files)}** файлов в папке `in/`")
    if in_files:
        log_table(["Файл"], [[Path(fp).name] for fp in in_files])
        log_md("", "INFO")  # Пустая строка для разделения абзацев

    # Подготовка skipped.xlsx и errors.xlsx
    skipped_wb = Workbook()
    # удалим стандартный лист (в openpyxl он создаётся по умолчанию)
    if skipped_wb.active:
        skipped_wb.remove(skipped_wb.active)

    errors_wb = Workbook()
    errors_ws = errors_wb.active
    errors_ws.append(["Файл", "Строка", "Описание ошибки"])

    total_filtered = 0
    total_skipped = 0
    total_errors = 0

    # Обработка каждого файла
    for fp in in_files:
        log_header(f"Обработка файла: {Path(fp).name}", 2)
        f, s, e = write_filtered_rows(fp, le_set, skipped_wb, errors_ws)
        total_filtered += f
        total_skipped += s
        total_errors += e
        log_file_separator()  # Добавляем разделитель между файлами

    # Сохраняем skipped.xlsx если есть листы
    if skipped_wb.sheetnames:
        try:
            skipped_out_path = os.path.join(OUT_DIR, SKIPPED_FILE)
            skipped_wb.save(skipped_out_path)
            logger.info("Сохранён файл с пропущенными строками: %s", skipped_out_path)
        except Exception as e:
            logger.exception("Ошибка при сохранении skipped.xlsx: %s", e)

    # Сохраняем errors.xlsx если были ошибки
    if total_errors > 0:
        try:
            errors_out_path = os.path.join(OUT_DIR, ERRORS_FILE)
            errors_wb.save(errors_out_path)
            logger.info("Сохранён файл ошибок: %s", errors_out_path)
        except Exception as e:
            logger.exception("Ошибка при сохранении errors.xlsx: %s", e)
    else:
        logger.info("Ошибок не обнаружено. Файл %s не создан.", ERRORS_FILE)

    # Финальный лог и вывод в консоль статистики
    log_header("Обработка завершена", 2)
    log_table(["Показатель", "Количество"],
                [["Отфильтровано", total_filtered],
                 ["Пропущено", total_skipped],
                 ["Ошибок", total_errors]])
    log_md("", "INFO")  # Пустая строка для разделения абзацев
    log_md("---", "INFO")  # Разделитель

    # Итоговая статистика в консоли
    log_total_stats(total_filtered, total_skipped, total_errors)

if __name__ == "__main__":
    main()

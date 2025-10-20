import os
from openpyxl import Workbook

# Папка для тестовых файлов
IN_DIR = "in"

# Создаём папку in, если её нет
os.makedirs(IN_DIR, exist_ok=True)

def create_empty_file(filename):
    """Создаёт файл с пустой таблицей данных."""
    wb = Workbook()
    ws = wb.active
    # Оставляем лист пустым
    wb.save(os.path.join(IN_DIR, filename))
    print(f"Создан {filename}: пустая таблица")

def create_no_columns_file(filename):
    """Создаёт файл без колонок Тип/Аналитика."""
    wb = Workbook()
    ws = wb.active
    # Добавляем заголовки без Тип и Аналитика
    ws['A1'] = 'Дата'
    ws['B1'] = 'Сумма'
    ws['C1'] = 'Описание'
    # Добавляем несколько строк данных
    ws['A2'] = '2023-01-01'
    ws['B2'] = 1000
    ws['C2'] = 'Тестовая проводка'
    wb.save(os.path.join(IN_DIR, filename))
    print(f"Создан {filename}: без колонок Тип/Аналитика")

def create_empty_le_file(filename):
    """Создаёт файл с пустыми значениями LE в строке."""
    wb = Workbook()
    ws = wb.active
    # Добавляем заголовки с Тип1 и Аналитика1
    ws['A1'] = 'ТИП1'
    ws['B1'] = 'АНАЛИТИКА1'
    ws['C1'] = 'Дата'
    ws['D1'] = 'Сумма'
    # Добавляем строки, где Тип=LE, но Аналитика пустая
    ws['A2'] = 'LE'
    ws['B2'] = ''  # Пустая Аналитика
    ws['C2'] = '2023-01-01'
    ws['D2'] = 1000
    # Ещё одна строка с данными
    ws['A3'] = 'ДРУГОЙ'
    ws['B3'] = 'TEST'
    ws['C3'] = '2023-01-02'
    ws['D3'] = 2000
    wb.save(os.path.join(IN_DIR, filename))
    print(f"Создан {filename}: с пустыми значениями LE")

def create_corrupted_file(filename):
    """Создаёт файл с поврежденными данными (просто пустой файл с расширением .xlsx для симуляции)."""
    # Для симуляции поврежденного файла, создаём пустой файл
    with open(os.path.join(IN_DIR, filename), 'w') as f:
        f.write("Это не .xlsx файл, симуляция повреждения")
    print(f"Создан {filename}: поврежденный файл (симуляция)")

def create_multiple_le_in_row_file(filename):
    """Создаёт файл с несколькими LE в одной строке."""
    wb = Workbook()
    ws = wb.active
    # Заголовки
    ws['A1'] = 'ТИП1'
    ws['B1'] = 'АНАЛИТИКА1'
    ws['C1'] = 'ТИП2'
    ws['D1'] = 'АНАЛИТИКА2'
    ws['E1'] = 'Дата'
    ws['F1'] = 'Сумма'
    # Строка с несколькими LE
    ws['A2'] = 'LE'
    ws['B2'] = 'TESTLE1'
    ws['C2'] = 'LE'
    ws['D2'] = 'TESTLE2'
    ws['E2'] = '2023-01-01'
    ws['F2'] = 1000
    # Другая строка
    ws['A3'] = 'ДРУГОЙ'
    ws['B3'] = 'TEST'
    ws['C3'] = 'LE'
    ws['D3'] = 'TESTLE3'
    ws['E3'] = '2023-01-02'
    ws['F3'] = 2000
    wb.save(os.path.join(IN_DIR, filename))
    print(f"Создан {filename}: с несколькими LE в одной строке")

def create_le_in_different_columns_file(filename):
    """Создаёт файл с LE в разных колонках."""
    wb = Workbook()
    ws = wb.active
    # Заголовки
    ws['A1'] = 'Дата'
    ws['B1'] = 'ТИП1'
    ws['C1'] = 'АНАЛИТИКА1'
    ws['D1'] = 'ТИП2'
    ws['E1'] = 'АНАЛИТИКА2'
    ws['F1'] = 'Сумма'
    # Строка с LE в колонке B
    ws['A2'] = '2023-01-01'
    ws['B2'] = 'LE'
    ws['C2'] = 'TESTLE1'
    ws['D2'] = 'ДРУГОЙ'
    ws['E2'] = 'TEST'
    ws['F2'] = 1000
    # Строка с LE в колонке D
    ws['A3'] = '2023-01-02'
    ws['B3'] = 'ДРУГОЙ'
    ws['C3'] = 'TEST'
    ws['D3'] = 'LE'
    ws['E3'] = 'TESTLE2'
    ws['F3'] = 2000
    wb.save(os.path.join(IN_DIR, filename))
    print(f"Создан {filename}: с LE в разных колонках")

def create_empty_rows_file(filename):
    """Создаёт файл с пустыми строками."""
    wb = Workbook()
    ws = wb.active
    # Заголовки
    ws['A1'] = 'ТИП1'
    ws['B1'] = 'АНАЛИТИКА1'
    ws['C1'] = 'Дата'
    ws['D1'] = 'Сумма'
    # Данные с пустыми строками
    ws['A2'] = 'LE'
    ws['B2'] = 'TESTLE1'
    ws['C2'] = '2023-01-01'
    ws['D2'] = 1000
    # Пустая строка 3
    # Данные в строке 4
    ws['A4'] = 'LE'
    ws['B4'] = 'TESTLE2'
    ws['C4'] = '2023-01-02'
    ws['D4'] = 2000
    wb.save(os.path.join(IN_DIR, filename))
    print(f"Создан {filename}: с пустыми строками")

def create_invalid_data_file(filename):
    """Создаёт файл с некорректными данными (LE без аналитики, неверные типы данных)."""
    wb = Workbook()
    ws = wb.active
    # Заголовки
    ws['A1'] = 'ТИП1'
    ws['B1'] = 'АНАЛИТИКА1'
    ws['C1'] = 'Дата'
    ws['D1'] = 'Сумма'
    # Строка с LE без аналитики
    ws['A2'] = 'LE'
    ws['B2'] = ''
    ws['C2'] = '2023-01-01'
    ws['D2'] = 1000
    # Строка с неверными данными
    ws['A3'] = 'LE'
    ws['B3'] = 'INVALIDLE'
    ws['C3'] = 'invalid_date'
    ws['D3'] = 'not_a_number'
    wb.save(os.path.join(IN_DIR, filename))
    print(f"Создан {filename}: с некорректными данными")

if __name__ == "__main__":
    create_empty_file("test_empty.xlsx")
    create_no_columns_file("test_no_columns.xlsx")
    create_empty_le_file("test_empty_le.xlsx")
    create_corrupted_file("test_corrupted.xlsx")
    create_multiple_le_in_row_file("test_multiple_le.xlsx")
    create_le_in_different_columns_file("test_le_different_columns.xlsx")
    create_empty_rows_file("test_empty_rows.xlsx")
    create_invalid_data_file("test_invalid_data.xlsx")
    print("Все тестовые файлы созданы в папке in/")